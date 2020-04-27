import datetime
import os
import string
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import Font
import win32com.client as win32

"""pyinstaller --onefile main.py """
# Компиляция в main.exe

number_of_Direction = {
    "Фонд / Фонд(рефинанс)": 0,
    "ЧИ (внутр)": 0
}
# Направления, дальше в них вкладываются стадии
Deal_stades = {
    "Всего поступило": 0,
    "Передано в Аналитический отдел": 0,
    "Вынесены на Кредитный комитет": 0,
    "Одобрено": 0,
    "Выдано": 0,
    "Основная": 0,
    "Доп.выдача": 0

}
# Стадии, подсчитывает в них и записывает с них в ячейки
my_columns = {
    "ID": 0,
    "Вид сделки": 0,
    "Стадия сделки": 0,
    "Направление сделки": 0,
    "Андеррайтер": 0,
    "Дата вынесения на КК": 0,
    "Результат (окончательно)": 0
}
# Названия колонок по которым идёт поиск

percent = " %"
chop = 0
direction_score = 1
bib = 1
other_Direction = 0
sheet = 0
workbook = 0
name_new_sheet = "Отчёт"
percent_Error = "0 %"
new_name_file = "report.xlsx"


def File_Converter_xls_to_xlsx():
    """Конвертирует файл report.xls, импортированный из Битрикс24 в *.xlsx для работы openpyxl"""
    global workbook
    global sheet
    global new_name_file

    file = str(os.path.abspath(os.curdir)) + "\\report.xlsx"
    file_old = str(os.path.abspath(os.curdir)) + "\\report.xls"

    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wbo = excel.Workbooks.Open(file_old)

    wbo.SaveAs(file, FileFormat=51)  # FileFormat = 51 is for .xlsx extension
    wbo.Close()  # FileFormat = 56 is for .xls extension
    del wbo
    excel.Application.Quit()

    workbook = openpyxl.load_workbook(filename=new_name_file)
    sheet = workbook.active

    Start()


def File_Saved():
    """Производит сохранение с заданным именем файла + время, и удалением промежуточного файла"""
    global workbook
    workbook.save("Отчёт. Статистика по Сделкам (" + modification_date(new_name_file) + ").xlsx")
    os.remove("report.xlsx")


def modification_date(filename):
    """Записывает дату после создания файла по его свойствам, по свежему так сказать"""
    t = os.path.getmtime(filename)
    return datetime.datetime.fromtimestamp(t).strftime("%d.%m")


def File_Format():
    """Форматирует ячейки листа после записи в файл"""
    for i in range(1, 31):
        sheet.row_dimensions[i].height = 20

    for c in string.ascii_letters:
        sheet.column_dimensions[c].width = 25
        # print(c)
    sheet.column_dimensions['A'].width = 46

    for r in range(1, 50):

        for w in range(1, 50):
            vino = sheet.cell(row=r, column=w)
            vino.alignment = Alignment(horizontal='center')


def File_List_Creator():
    """ Создаёт лист внутри нового файла для записи данных"""
    workbook.create_sheet(index=1, title=name_new_sheet)


def Convertering():
    """ Вкладывает словарь в поле значения другого словоря """

    for key in number_of_Direction:
        number_of_Direction[key] = dict(Deal_stades)


def Reading_Appraiser():
    """ Производит подсчёт общего числа сделок. Это требуется для корректной работы циклов """

    nuuumer = 0

    while True:
        nuuumer = nuuumer + 1
        this_sources = sheet.cell(row=nuuumer, column=my_columns["ID"]).value

        global direction_score

        if this_sources != sheet.cell(row=999, column=1).value:
            direction_score = direction_score + 1

        if this_sources == sheet.cell(row=999, column=1).value:
            break


def Automatic_search_of_the_columns():
    """ Находит названия столбцов по заданным именам из массива. Сохранет цифрой номер столбца"""

    num_columns_w = 0

    while True:
        num_columns_w = num_columns_w + 1
        this_columns = sheet.cell(row=1, column=num_columns_w).value

        if this_columns in my_columns:
            my_columns[this_columns] = num_columns_w
            # print("Успешно Найдено: "+this_columns + " - А её номер : = " + str(num_columns_w))

        if this_columns == sheet.cell(row=1, column=105).value:
            break
    Reading_Appraiser()


def Reading_WriterArray():
    """ Для ключевых полей производится подсчёт. Цикл проходит по всем сделкам считывая необходимые ячейки ключчевых столбцов
    Некторые поля подсчитывают по заполнению, а в ячейках других идёт поиск определённых слов
    """

    print("Metod Reading Writer Array: - Done")

    empty_cell = sheet.cell(row=999, column=1).value
    # Пустая ячейка как понятие истинной пустоты

    colum = 0
    pole = " "
    this_cell = 0

    for i in range(2, direction_score):

        direction_key = 0
        this_cell_Direct = sheet.cell(row=i, column=my_columns[
            "Направление сделки"]).value  # Подсчёт общего количества сделок по направлениям

        """Это разграничения данных и условий для этих данных"""

        if this_cell_Direct == "ФОНД" or this_cell_Direct == "ФОНД (рефинанс)":
            number_of_Direction["Фонд / Фонд(рефинанс)"]["Всего поступило"] += 1
            direction_key = "Фонд / Фонд(рефинанс)"
        elif this_cell_Direct == "ЧИ (внутр)":
            number_of_Direction["ЧИ (внутр)"]["Всего поступило"] += 1
            direction_key = "ЧИ (внутр)"

        colum = int(my_columns["Андеррайтер"])
        this_cell = sheet.cell(row=i, column=colum).value

        if this_cell != empty_cell:
            pole = "Передано в Аналитический отдел"
            number_of_Direction[direction_key][pole] += 1

        colum = int(my_columns["Дата вынесения на КК"])
        this_cell = sheet.cell(row=i, column=colum).value
        if this_cell != empty_cell:
            number_of_Direction[direction_key]["Вынесены на Кредитный комитет"] += 1

        colum = int(my_columns["Результат (окончательно)"])
        this_cell = sheet.cell(row=i, column=colum).value
        this_cell = str(this_cell)
        this_cell = this_cell.lower()
        if this_cell == "одобрено" or this_cell == "положительно" or this_cell == "одобрили":
            number_of_Direction[direction_key]["Одобрено"] += 1

        colum = int(my_columns["Стадия сделки"])
        this_cell = sheet.cell(row=i, column=colum).value
        this_cell = str(this_cell)
        this_cell = this_cell.lower()

        if this_cell == "обслуживание" or this_cell == "просрочка" or this_cell == "сделка успешно закрыта":
            number_of_Direction[direction_key]["Выдано"] += 1
            if sheet.cell(row=i, column=int(my_columns["Вид сделки"])).value == "основная":
                number_of_Direction[direction_key]["Основная"] += 1
            elif sheet.cell(row=i, column=int(my_columns["Вид сделки"])).value == "доп. выдача":
                number_of_Direction[direction_key]["Доп.выдача"] += 1


def WriterFile():
    """ Запись Содержимого словарей в файл, расчёт % """

    global bib
    sheet.cell(row=1, column=1).value = "Наименование направлений"
    sheet.cell(row=1, column=1).font = Font(bold=True)
    column_num_write = 2

    save_tottal_cell = 0

    save_tottal_cell_odobreno = 0

    for num_stades in Deal_stades:
        bib = bib + 1
        sheet.cell(row=bib, column=1).value = num_stades

    for direction in number_of_Direction:

        num = 1
        sheet.cell(row=1, column=column_num_write).value = direction
        sheet.cell(row=1, column=column_num_write).font = Font(bold=True)

        sheet.cell(row=1, column=column_num_write + 1).value = "% от входящих"
        sheet.cell(row=1, column=column_num_write + 1).font = Font(bold=True)

        sheet.cell(row=1, column=column_num_write + 2).value = "% выдачи из одобренных"
        sheet.cell(row=1, column=column_num_write + 2).font = Font(bold=True)

        for dil in number_of_Direction[direction]:
            num = num + 1

            sheet.cell(row=num, column=column_num_write).value = number_of_Direction[direction][dil]
            if num == 2:
                save_tottal_cell = number_of_Direction[direction][dil]
            if num > 2 and num < 6:
                del_a = int(number_of_Direction[direction][dil])
                del_b = save_tottal_cell

                if del_a > 0 and del_b > 0:
                    drop = del_a / del_b * 100
                    drop = str(drop)
                    drop_t = int(drop.find('.') + 2)
                    drops = drop[0:drop_t]
                    sheet.cell(row=num, column=column_num_write + 1).value = str(drops) + percent
                else:
                    sheet.cell(row=num, column=column_num_write + 1).value = percent_Error
            elif num > 6:
                del_a = int(number_of_Direction[direction][dil])
                del_b = int(number_of_Direction[direction]["Выдано"])

                if del_a > 0 and del_b > 0:
                    drop = del_a / del_b * 100
                    drop = str(drop)
                    drop_t = int(drop.find('.') + 2)
                    drops = drop[0:drop_t]
                    sheet.cell(row=num, column=column_num_write + 1).value = str(drops) + percent
                else:
                    sheet.cell(row=num, column=column_num_write + 1).value = percent_Error

            if dil == "Одобрено":
                save_tottal_cell_odobreno = number_of_Direction[direction][dil]

            if dil == "Выдано":
                del_C = int(number_of_Direction[direction][dil])
                del_V = save_tottal_cell_odobreno

                if del_C > 0 and del_V > 0:
                    drop = del_C / del_V * 100
                    drop = str(drop)
                    drop_t = int(drop.find('.') + 2)
                    drops = drop[0:drop_t]
                    sheet.cell(row=num, column=column_num_write + 2).value = str(drops) + percent
                else:
                    sheet.cell(row=num, column=column_num_write + 2).value = percent_Error

        column_num_write += 3


def Start():
    """ Скрипт последовательности работы методов. Используется для управления """

    global sheet
    Automatic_search_of_the_columns()
    Convertering()
    Reading_WriterArray()
    File_List_Creator()
    sheet = workbook[name_new_sheet]
    File_Format()
    WriterFile()
    File_Saved()


for i in range(1, 2):
    File_Converter_xls_to_xlsx()
